import math
import pandas as pd
import mysql.connector
from openpyxl import load_workbook
import numpy as np
import py_to_sql
import glob
import time
from datetime import datetime

# useful functions

# returns a list with duplicate columns name
def find_duplicates(dataframe):
  dup = dataframe.columns.value_counts() > 1
  return list(dup[dup].index)

# returns the position of last duplicate column
def find_last_dup_pos(value, list):
  list.reverse()
  pos = len(list) - 1 - list.index(value)
  return pos

# connect to MySQL
mydb = mysql.connector.connect(
  host = "localhost",
  user = "root",
  password ="dimosk",
  database = "stocks",
  auth_plugin = "mysql_native_password"
)

# iteration counter
cnt_t = 0

dataframes_list = []

for file in glob.glob("us_data/*.xlsx") :

  # creating a list with the names of all sheets in an Excel file
  excel_sheets_names = load_workbook(file, read_only = True, keep_links=False).sheetnames

  # define rows that need to be deleted in every sheet
  rows_to_del = ['Income Statement', 'Balance Sheet', 'Cash Flow Statement', 'Non-GAAP Metrics', 'Valuation Measures',
                'Valuation Ratios', 'Liquidity/Efficiency Ratios', 'Profitability Ratios', 'Return Ratios']



  # loop that loads excel sheet one by one
  for cnt in range(len(excel_sheets_names)):

    data = pd.read_excel(file, sheet_name = excel_sheets_names[cnt] ,header = 2, index_col = 0 )
    data = data.drop(rows_to_del)


  # rename rows
  # e.g if the month is March(3) or April(4) we consider that the financial statements correspond to Q1 results.
  # We work with quarterly statements, therefore we have four different report periods (Q1,Q2,Q3,Q4)
    col_names = []
    for i in data.columns :
      if not isinstance(i, (pd.Timestamp, datetime)):
        data.drop(i, axis=1, inplace=True)
        continue
      if str(i)[6] == "3" or str(i)[6] == "4" :
        t = "Q1-" + str(i)[0:4]
        col_names.append(t)
      elif str(i)[6] == "6" or str(i)[6] == "7" :
        t = "Q2-" + str(i)[0:4]
        col_names.append(t)
      elif str(i)[6] == "9" or str(i)[5:7] == "10" :
        t = "Q3-" + str(i)[0:4]
        col_names.append(t)
      elif str(i)[5:7] == "12" :
        t = "Q4-" + str(i)[0:4]
        col_names.append(t)
      elif str(i)[6] == "1":
        t = "Q4-" + str(int(str(i)[0:4]) -1)
        col_names.append(t)
      else :
        data.drop(i, axis = 1, inplace = True)

    data.columns = col_names

    # transpose dataframe
    data = data.transpose()
    #add column with company name
    data.insert(0, "Ticker", excel_sheets_names[cnt])

    # Duplicate columns refer to different items, so they should not be deleted
    pos_gross = find_last_dup_pos("Gross Margin", list(data.columns))
    pos_net = find_last_dup_pos("Net Income", list(data.columns))
    pos_inv = find_last_dup_pos("Inventory", list(data.columns))
    net_income = data["Net Income"].iloc[:,1]
    inventory = data["Inventory"].iloc[:,1]
    gross_margin = data["Gross Margin"].iloc[:,1]
    data = data.loc[:,~data.columns.duplicated()]
    data.insert(pos_net, "Net Income (Cash Flow)", net_income)
    data.insert(pos_inv, "Change of Inventory", inventory)
    data.insert(pos_inv, "Gross Margin Ratio", inventory)

    #converting index to column and rename it
    data.reset_index(inplace = True)
    data.rename(columns = { "index" : "Report Date"}, inplace = True)
    
    dataframes_list.append(data)

final_data = pd.concat(dataframes_list)

# create table "fin_stat"

pySQL_functions.create_table(final_data, "fin_stat", mydb)
pySQL_functions.insert_to_sql(final_data, "fin_stat", mydb)


# create table "comp_info"

data = pd.read_excel("us_data_sample/us_listed_sample.xlsx", header = 2)
data = data.rename(columns = lambda x : x.strip())
data.drop(index = 0, inplace = True)
data.drop(labels = "Description", axis = 1,inplace = True)

pySQL_functions.create_table(data, "comp_info", mydb)
pySQL_functions.insert_to_sql(data, "comp_info", mydb)


# create table with daily stock prices from Marketstack, table "daily_prices"

data_vendor = "Marketstack"
mycursor = mydb.cursor()
query_create = "CREATE TABLE IF NOT EXISTS daily_prices( " \
               "`Ticker` char(10), `Date` date, ` Adj_Close Price` decimal(15,3), `Close Price` decimal(15,3), " \
               "`Split Factor` float, `Dividend` float, `Data Vendor` char(25))"
mycursor.execute(query_create)
mydb.commit()

# load comp_info table from SQL database into a list (select only column "Ticker")
mycursor.execute("select `Ticker` from companies")
data_sql = mycursor.fetchall()

# retrieve historical daily prices for every company that exists in table "comp_info"
# API response can return up to 1000 prices, therefore for some companies I need to send multiple API requests 

data_vendor = "Marketstack"

for x in data_sql:
    end_date = datetime.now().date()
    while True :
        query = {"symbols": x[0], "date_to": end_date, "limit": 1000}
        url = "http://api.marketstack.com/v1/eod?access_key=**********"
        api_response = requests.get(url, params=query)
        response_json = json.loads(api_response.text)
        if api_response.status_code == 422 or not response_json["data"]:
            break
        else :
            for i in response_json["data"]:
                query_ins = "INSERT INTO daily_prices VALUES (DEFAULT,%s, %s, %s, %s, %s, %s, %s)"
                val = (x[0], i["date"].split("T")[0], i["adj_close"], i["close"], i["split_factor"], i["dividend"], data_vendor)
                mycursor.execute(query_ins, val)
                mydb.commit()
            end_date = datetime.strptime(response_json["data"][-1]["date"], "%Y-%m-%dT%H:%M:%S+%f") - timedelta(1)
            if len(response_json["data"]) < 1000 :
                break
           
